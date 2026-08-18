"""
exporter.py
===========
Multi-sheet Excel workbook generator using OpenPyXL, producing:
  1. Summary_Dashboard      - high-level KPIs and monthly staffing budget
  2. 30Min_Interval_Plan    - granular interval forecast + staffing + SLA
  3. Model_Leaderboard      - backtesting evaluation metrics for all models
"""

from __future__ import annotations

import io
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
KPI_LABEL_FONT = Font(bold=True, size=11)
KPI_VALUE_FONT = Font(size=12, color="1F4E78", bold=True)


def _style_header_row(ws, row_idx: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(ws, df: pd.DataFrame, start_row: int = 1):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:200]]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 40)


def _write_dataframe(ws, df: pd.DataFrame, start_row: int = 1):
    for j, col_name in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=str(col_name))
    _style_header_row(ws, start_row, len(df.columns))

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            if pd.isna(val):
                val = None
            elif hasattr(val, "isoformat"):
                val = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
            ws.cell(row=i, column=j, value=val)

    _autofit_columns(ws, df, start_row)


def _write_summary_dashboard(wb: Workbook, kpis: dict, monthly_budget_df: Optional[pd.DataFrame]):
    ws = wb.active
    ws.title = "Summary_Dashboard"

    ws["A1"] = "PyErlang-ML — Enterprise Capacity Plan Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    row = 3
    for label, value in kpis.items():
        ws.cell(row=row, column=1, value=label).font = KPI_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = KPI_VALUE_FONT
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    if monthly_budget_df is not None and not monthly_budget_df.empty:
        budget_start = row + 2
        ws.cell(row=budget_start, column=1, value="Monthly Staffing Budget").font = TITLE_FONT
        _write_dataframe(ws, monthly_budget_df, start_row=budget_start + 1)


def _write_interval_plan(wb: Workbook, interval_plan_df: pd.DataFrame):
    ws = wb.create_sheet("30Min_Interval_Plan")
    _write_dataframe(ws, interval_plan_df, start_row=1)

    # Add a simple staffing vs volume line chart if enough rows exist
    if len(interval_plan_df) >= 2 and {"forecast_volume", "gross_fte"}.issubset(interval_plan_df.columns):
        try:
            chart = LineChart()
            chart.title = "Forecast Volume vs Gross FTE"
            chart.y_axis.title = "Value"
            chart.x_axis.title = "Interval"

            vol_col = interval_plan_df.columns.get_loc("forecast_volume") + 1
            fte_col = interval_plan_df.columns.get_loc("gross_fte") + 1
            max_row = len(interval_plan_df) + 1

            vol_ref = Reference(ws, min_col=vol_col, min_row=1, max_row=max_row)
            fte_ref = Reference(ws, min_col=fte_col, min_row=1, max_row=max_row)
            chart.add_data(vol_ref, titles_from_data=True)
            chart.add_data(fte_ref, titles_from_data=True)
            ws.add_chart(chart, f"{get_column_letter(len(interval_plan_df.columns) + 2)}2")
        except Exception:
            pass  # chart is a nice-to-have; never fail the export because of it


def _write_model_leaderboard(wb: Workbook, leaderboard_df: pd.DataFrame):
    ws = wb.create_sheet("Model_Leaderboard")
    if leaderboard_df is None or leaderboard_df.empty:
        ws["A1"] = "No backtesting results available."
        return
    _write_dataframe(ws, leaderboard_df, start_row=1)


def build_excel_workbook(
    kpis: dict,
    interval_plan_df: pd.DataFrame,
    leaderboard_df: Optional[pd.DataFrame] = None,
    monthly_budget_df: Optional[pd.DataFrame] = None,
) -> bytes:
    """
    Assemble the full multi-sheet workbook in memory and return raw bytes,
    suitable for st.download_button(data=...).
    """
    wb = Workbook()
    _write_summary_dashboard(wb, kpis, monthly_budget_df)
    _write_interval_plan(wb, interval_plan_df)
    _write_model_leaderboard(wb, leaderboard_df if leaderboard_df is not None else pd.DataFrame())

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
